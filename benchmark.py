#!/usr/bin/env python3

import argparse
import os, sys
import shutil
import subprocess
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime
import traceback


parser = argparse.ArgumentParser(description="Benchmark LLVM scheduler configurations with GadgetSetAnalyzer.")
parser.add_argument("-b", "--benchmarks", default="all", help="Benchmarks to run (comma-separated list)")
parser.add_argument("--flags", default="", help="Clang flags to use when compiling benchmarks")
parser.add_argument("-d", "--debug", action="store_true", help="Show command output")
parser.add_argument("--skip-build", action="store_true", help="Skip the build step (just compare and generate the results spreadsheet)")
parser.add_argument("--skip-compare", action="store_true", help="Skip the compare step (just generate the results spreadsheet)")
args = parser.parse_args()


configs = {
    "Original": "--misched=default --misched-postra=false",
    "Pre-RA": "--misched=ropsched --misched-postra=false",
    "Post-RA": "--misched=default --misched-postra=true",
    "Both": "--misched=ropsched --misched-postra=true",
}

cwd = Path(os.path.realpath(__file__)).expanduser().resolve().parent  # Everything is relative to where this script is actually located.
clang = cwd / "llvm-ropsched/build/bin/clang"
gsa = cwd / "GadgetSetAnalyzer"
results = cwd / "results"
binaries = results / "binaries"
output = results / "results.xlsx"
time = datetime.now().strftime("%b %d, %Y %I∶%M∶%S %p")

benchmarks = cwd / "benchmarks"
benchmarks = list(benchmarks.iterdir())


if args.benchmarks != "all":
    selected = args.benchmarks.split(",")
    available = [benchmark.name for benchmark in benchmarks]

    for benchmark in selected:
        if benchmark not in available:
            parser.error(f"unknown benchmark '{benchmark}'")

    benchmarks = sorted(list(filter(lambda benchmark: benchmark.name in set(selected), benchmarks)))


def format(x: float) -> str:
    sign = "+" if x > 0 else ""
    return f"{sign}{x:.3f}"


def highlight(ws, cell) -> None:
    header = ws.cell(1, cell.column).value
    value = str(cell.value)

    if ("+" in value and "Gadget Quality" in header) or ("-" in value and "Number" in header):
        cell.style = "Good"
    elif ("-" in value and "Gadget Quality" in header) or ("+" in value and "Number" in header):
        cell.style = "Bad"
    elif "(" in value:
        cell.style = "Neutral"


def run(cmd: str, cwd: Path, env: dict = None) -> None:
    stdout, stderr = (None, None) if args.debug else (subprocess.DEVNULL, subprocess.DEVNULL)
    subprocess.run(cmd, cwd=cwd, env=env, stdout=stdout, stderr=stderr, check=True)


def build(benchmark: Path) -> list[Path]:
    release = benchmark / "target" / "release"
    paths = []

    for config, misched in configs.items():
        if os.path.exists(release):
            shutil.rmtree(release)

        flags = misched.split(" ") + args.flags.split(" ")
        flags = [f"-C llvm-args={flag}" for flag in flags]
        flags = " ".join(flags)
        env = os.environ.copy()
        env["RUSTFLAGS"] = flags

        run(["cargo", "build", "--release"], benchmark, env=env)

        binary = None

        for filename in sorted(os.listdir(release)):
            filepath = release / filename
            if os.path.isfile(filepath) and os.access(filepath, os.X_OK):
                binary = filepath
                break

        if binary is None:
            for filename in sorted(os.listdir(release)):
                if filename.endswith(".rlib"):
                    filepath = release / filename
                    binary = filepath
                    break

        if binary is None:
            return paths

        filename = f"{benchmark.name}.{config}" if (benchmark.name == binary.name) else f"{benchmark.name} ({filename}).{config}"
        destination = binaries / filename

        shutil.copy2(release / binary, destination)

        paths.append(destination)

    return paths


def compare(original: Path, variants: list[Path]) -> Path:
    name = original.stem
    old = gsa / "results" / name
    variants = [f"{variant.suffix[1:]}={variant}" for variant in variants]

    if not args.skip_compare:
        if os.path.exists(old):
            shutil.rmtree(old)

        run(["python3", "src/GSA.py", str(original), "--variants", *variants, "--output_metrics", "--result_folder_name", name], cwd=gsa)

    return old / "Gadget Quality.csv"


def combine(files: list[Path]):
    dataframes = []

    # Combine all the individual CSV files into a single dataframe.
    for csv in files:
        df = pd.read_csv(csv)
        df.insert(0, "Benchmark", csv.parent.name)
        dataframes.append(df)

    data = pd.concat(dataframes, ignore_index=True)

    # Convert the dataframe to an Excel workbook.
    mode = "w"
    if_new_sheet_exists = None

    # Add to the existing results if the spreadsheet already exists.
    if output.exists():
        mode = "a"
        if_new_sheet_exists = "new"

    with pd.ExcelWriter(output, engine="openpyxl", mode=mode, if_sheet_exists=if_new_sheet_exists) as writer:
        data.to_excel(writer, sheet_name=time, index=False)

    # Format the results to be a little prettier.
    wb = load_workbook(output)
    ws = wb[wb.sheetnames[-1]]

    # Start by merging the benchmark names column.
    for row in range(2, ws.max_row, 4):
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=1)
        ws.cell(row, 1).alignment = Alignment(vertical="center", horizontal="center")

    # Highlight the results based on whether it was positive, negative, or neutral.
    for row in ws.iter_cols(3):
        for cell in filter(lambda cell: cell.value is not None, row[1:]):
            highlight(ws, cell)

    # Calculate the average results for each category accross each variant.
    averages = [[0 for _ in range(6)] for _ in range(3)]  # 6 categories of ROP metrics x 3 scheduling variants.

    for benchmark in range(len(benchmarks)):
        srow = benchmark * 4 + 2

        for i, row in enumerate(ws.iter_rows(srow + 1, srow + 4 - 1)):
            for j, cell in enumerate(row[2:]):
                difference = cell.value[cell.value.find('(') + 1: cell.value.find(')')]
                averages[i][j] += float(difference)

    # Write the flags for the run.
    lrow = ws.max_row + 2
    ws.merge_cells(f"B{lrow}:I{lrow}")
    ws[f"A{lrow}"].value = "Extra Flags"
    ws[f"B{lrow}"].value = args.flags if args.flags else "None"

    # Write the average results for the run.
    lrow += 1
    ws.merge_cells(start_row=lrow, start_column=1, end_row=lrow + 2, end_column=1)
    cell = ws.cell(lrow, 1, "Average Difference")
    cell.alignment = Alignment(vertical="center", horizontal="center")

    for i, config in enumerate(list(configs.keys())[1:]):
        ws.cell(lrow + i, 2, config)

    for i in range(len(averages)):
        for j in range(len(averages[0])):
            averages[i][j] /= len(benchmarks)
            cell = ws.cell(lrow + i, 3 + j, value=format(averages[i][j]))
            highlight(ws, cell)

    wb.save(output)


if not os.path.exists(binaries):
    os.makedirs(binaries)

try:
    files = []

    for benchmark in benchmarks:
        try:
            paths = build(benchmark)
            path = compare(paths[0], paths[1:])
            files.append(path)
        except Exception as e:
            print("Failed: ", benchmark, e)
            continue

    path = combine(files)

except Exception as exception:
    if args.debug:
        traceback.print_exc()
    else:
        _, _, tb = sys.exc_info()
        tb = traceback.extract_tb(tb)[-1]
        print(f"An error occurred: {exception}")
